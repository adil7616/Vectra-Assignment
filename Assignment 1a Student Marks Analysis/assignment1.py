# -*- coding: utf-8 -*-
"""
Created on Tue Sep  2 08:21:53 2025

@author: adilm
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference

# Load Excel data (student.xlsx)
df = pd.read_excel("student.xlsx")

# Subjects list
subjects = ["Math", "Physics", "Chemistry", "Biology"]

# Vectorized Computations
df["Total"] = df[subjects].sum(axis=1)
df["Average"] = df["Total"] / len(subjects)

# Assign Grades using vectorized conditions
conditions = [
    (df["Average"] >= 90),
    (df["Average"] >= 75) & (df["Average"] < 90),
    (df["Average"] >= 60) & (df["Average"] < 75),
    (df["Average"] < 60)
]

grades = ["A", "B", "C", "F"]
df["Grade"] = np.select(conditions, grades, default="F")

# Find top performers per subject
top_performers = {}
for subject in subjects:
    top_performers[subject] = df.nlargest(3, subject)[["StudentID", "Name", subject]]

# Convert top performers dictionary to DataFrame for saving
top_df = pd.concat(top_performers, axis=0)

# Save to Excel with 2 sheets
with pd.ExcelWriter("results.xlsx", engine="openpyxl") as writer:
    df[["StudentID", "Name", "Total", "Average", "Grade"]].to_excel(writer, sheet_name="Summary", index=False)
    top_df.to_excel(writer, sheet_name="Top Performers")

# Open the results.xlsx workbook for chart creation
wb = load_workbook("results.xlsx")
ws = wb["Summary"]

# Create Bar Chart for Average per Subject
avg_per_subject = df[subjects].mean()
chart_df = avg_per_subject.reset_index()
chart_df.columns = ["Subject", "Average Marks"]

# Add chart data to a new sheet
ws_chart = wb.create_sheet("Charts")
for r in dataframe_to_rows(chart_df, index=False, header=True):
    ws_chart.append(r)

# Create bar chart
chart = BarChart()
chart.title = "Average Marks per Subject"
chart.x_axis.title = "Subjects"
chart.y_axis.title = "Average Marks"

data = Reference(ws_chart, min_col=2, min_row=1, max_row=len(chart_df) + 1)
cats = Reference(ws_chart, min_col=1, min_row=2, max_row=len(chart_df) + 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws_chart.add_chart(chart, "E5")

# Save workbook again
wb.save("results.xlsx")
